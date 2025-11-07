import pandas as pd
import os
from openpyxl import Workbook

# ---------------------------------------------------------
# ✅ LIMPAR COLUNAS UNNAMED
# ---------------------------------------------------------
def limpar_unnamed(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.str.contains("^Unnamed")]


# ---------------------------------------------------------
# ✅ CARREGAR SIGA
# ---------------------------------------------------------
def carregar_csv_siga(arquivo):
    df = pd.read_csv(arquivo, sep=",", encoding="latin-1", dtype=str)
    df = limpar_unnamed(df)

    # Padroniza colunas importantes
    df.rename(columns={
        "Código": "Código",
        "Nome": "Nome",
        "Dependência": "Dependência"
    }, inplace=True)

    return df


# ---------------------------------------------------------
# ✅ CARREGAR FORMULÁRIO
# ---------------------------------------------------------
def carregar_csv_form(arquivo):
    df = pd.read_csv(arquivo, sep=",", encoding="latin-1", dtype=str)
    df = limpar_unnamed(df)

    # Garantir que exista coluna Observações
    if "Observações" not in df.columns:
        df["Observações"] = ""

    df.rename(columns={
        "Nome / Tipo de Bens": "Nome",
        "Dependência / Localização": "Dependência"
    }, inplace=True)

    return df


# ---------------------------------------------------------
# ✅ GERAR CÓDIGO ÚNICO AUTOMÁTICO PARA O FORMULÁRIO
# ---------------------------------------------------------
def gerar_codigo_unico(indice: int) -> str:
    return f"FORM-{indice+1:05d}"


# ---------------------------------------------------------
# ✅ PREPARAR EXIBIÇÃO VISUAL (NÃO USADO NA EXPORTAÇÃO)
# ---------------------------------------------------------
def preparar_dados_para_visual(df_siga, df_form):
    df_siga["nome_visual"] = df_siga["Nome"]
    df_form["nome_visual"] = df_form["Nome"] + " — " + df_form["Observações"].fillna("")
    return df_siga, df_form


# ---------------------------------------------------------
# ✅ GERAR EXCEL COMPLETO (Várias abas)
# ---------------------------------------------------------
def gerar_excel_completo(df_siga, df_form, pareamentos):
    caminho = "comparacao_completa.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pareamentos"

    ws1.append(["Código SIGA", "Nome SIGA", "Código Formulário", "Nome Formulário", "Observações"])

    for idx, escolha in pareamentos.items():
        if escolha == "(Nenhum)":
            ws1.append([
                df_siga.loc[idx, "Código"],
                df_siga.loc[idx, "Nome"],
                "",
                "",
                ""
            ])
        else:
            codigo_form = escolha.split("|")[0].strip()
            linha_form = df_form[df_form["codigo_formulario"] == codigo_form].iloc[0]

            ws1.append([
                df_siga.loc[idx, "Código"],
                df_siga.loc[idx, "Nome"],
                linha_form["codigo_formulario"],
                linha_form["Nome"],
                linha_form["Observações"]
            ])

    # Aba SIGA
    ws2 = wb.create_sheet("SIGA")
    ws2.append(list(df_siga.columns))
    for _, r in df_siga.iterrows():
        ws2.append(list(r.values))

    # Aba Formulário
    ws3 = wb.create_sheet("Formulário")
    ws3.append(list(df_form.columns))
    for _, r in df_form.iterrows():
        ws3.append(list(r.values))

    wb.save(caminho)
    return caminho


# ---------------------------------------------------------
# ✅ GERAR CSV COMPLETO (Plano, tudo junto)
# ---------------------------------------------------------
def gerar_csv_completo(df_siga, df_form, pareamentos):
    linhas = []

    for idx, escolha in pareamentos.items():
        d = {
            "SIGA__Codigo": df_siga.loc[idx, "Código"],
            "SIGA__Nome": df_siga.loc[idx, "Nome"],
            "SIGA__Dependencia": df_siga.loc[idx, "Dependência"],
        }

        if escolha == "(Nenhum)":
            # Sem correspondente
            d.update({
                "FORM__Codigo": "",
                "FORM__Nome": "",
                "FORM__Observacoes": "",
                "FORM__Dependencia": "",
            })
        else:
            codigo_form = escolha.split("|")[0].strip()
            linha_form = df_form[df_form["codigo_formulario"] == codigo_form].iloc[0]

            d.update({
                "FORM__Codigo": linha_form["codigo_formulario"],
                "FORM__Nome": linha_form["Nome"],
                "FORM__Observacoes": linha_form["Observações"],
                "FORM__Dependencia": linha_form["Dependência"],
            })

        linhas.append(d)

    df_final = pd.DataFrame(linhas)
    caminho = "comparacao_completa.csv"
    df_final.to_csv(caminho, index=False, encoding="utf-8-sig")

    return caminho
